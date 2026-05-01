# Este script cria uma planilha Excel estruturada para controle de Reels da Legacy Academia
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Criar o arquivo Excel
wb = Workbook()

# -----------------
# Aba 1 - Produção de Reels
# -----------------
sheet1 = wb.active
sheet1.title = 'Producao de Reels'
headers = [
    'Data', 'Tema / Pilar', 'Objetivo (Dor / Desejo / Objeção / Prova)',
    'Mensagem Principal', 'CTA', 'Status (Gravado / Editado / Postado)',
    'Link do Reels', 'Visualizações', 'Comentários', 'Leads (Directs)'
]
sheet1.append(headers)

# Estilo cabeçalho
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
header_font = Font(color='FFFFFF', bold=True)
for col in range(1, len(headers)+1):
    c = sheet1.cell(row=1, column=col)
    c.fill = header_fill
    c.font = header_font
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Linhas exemplo
example_rows = [
    ['13/10', 'Ativação Metabólica Progressiva', 'Autoridade / Ciência', 'Teu corpo continua queimando gordura até 36h…', 'Digite CIÊNCIA', '', '', '', '', ''],
    ['14/10', 'Metabolismo travado', 'Dor', 'O erro mais comum é comer de menos.', 'Digite CIÊNCIA', '', '', '', '', ''],
]
for row in example_rows:
    sheet1.append(row)

# Ajustar largura
for col in sheet1.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    sheet1.column_dimensions[column].width = adjusted_width

# -----------------
# Aba 2 - Resumo de Performance
# -----------------
sheet2 = wb.create_sheet('Resumo de Performance')
headers2 = ['Métrica', 'Semana 1', 'Semana 2', 'Semana 3', 'Semana 4', 'Média Geral']
sheet2.append(headers2)
metrics = [
    'Média de visualizações',
    'Média de comentários “Ciência”',
    'Leads (Directs recebidos)',
    'Conversões WhatsApp',
    'Vendas (Emagreça em Casa)'
]
for m in metrics:
    sheet2.append([m, '', '', '', '', '=AVERAGE(B{row}:E{row})'.format(row=sheet2.max_row+1)])

for col in range(1, len(headers2)+1):
    c = sheet2.cell(row=1, column=col)
    c.fill = header_fill
    c.font = header_font
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# -----------------
# Aba 3 - Calendário de Conteúdo
# -----------------
sheet3 = wb.create_sheet('Calendario de Conteudo')
headers3 = ['Data', 'Tema', 'Tipo de Conteúdo', 'Status', 'Link']
sheet3.append(headers3)
for col in range(1, len(headers3)+1):
    c = sheet3.cell(row=1, column=col)
    c.fill = header_fill
    c.font = header_font
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

example_calendar = [
    ['13/10', 'Ativação Metabólica Progressiva', 'Reels', '✅ Postado', ''],
    ['14/10', 'Metabolismo travado', 'Reels', '🟨 Gravado', ''],
]
for row in example_calendar:
    sheet3.append(row)

# Salvar arquivo
wb.save('Planilha_Producao_Reels_Legacy.xlsx')
